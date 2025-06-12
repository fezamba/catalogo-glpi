import pandas as pd
import mysql.connector
from datetime import datetime
import openpyxl
import re
import warnings
from openpyxl.utils import get_column_letter

warnings.simplefilter("ignore")

excel_path = 'c:/Users/fpetit/Desktop/workspace/PHP-CATALOGO/bd_script/script_pandas/catalogo.xlsx'
abas = [
    'Credenciais de Acesso',
    'Segurança',
    'Infraestrutura',
    'Microsoft 365',
    'Sistemas e Portais',
    'Softwares Homologados'
]

conn = mysql.connector.connect(
    host='127.0.0.1',
    user='root',
    password='sefazfer123@',
    database='catalogo-teste'
)
cursor = conn.cursor()
agora = datetime.now()

wb_links = openpyxl.load_workbook(excel_path, data_only=True)

servicos_inseridos = {}

for aba in abas:
    print(f"Processando aba: {aba}")
    df = pd.read_excel(excel_path, sheet_name=aba)
    df.columns = df.columns.str.strip()
    ws = wb_links[aba]

    coluna_subcat = next((col for col in df.columns if 'sub' in col.lower()), None)
    if not coluna_subcat:
        print(f"[ERRO] Coluna de subcategoria não encontrada na aba '{aba}'")
        continue

    df[coluna_subcat] = df[coluna_subcat].ffill()
    df = df.ffill()

    for i, row in df.iterrows():
        nome_servico = row.get('Serviço') or row.get('Nome do Serviço') or row.get('Título')
        if pd.isna(nome_servico): continue

        nome_subcategoria = row.get(coluna_subcat, '')
        nome_subcategoria = str(nome_subcategoria).strip()
        if not nome_subcategoria:
            continue
        nome_subcategoria = re.sub(r'\s+', ' ', nome_subcategoria)

        descricao = row.get('Descrição do Serviço')
        descricao = descricao if pd.notna(descricao) and str(descricao).strip() else "Descrição padrão."

        # Avaliar a necessidade desse bloco
        
        # descricaotec = row.get('Descrição Técnica')
        # descricaotec = descricaotec if pd.notna(descricaotec) else ""
        # descricoes_lista = [d.strip() for d in str(descricaotec).split('\n') if d.strip()]
        
        area = row.get('Equipe Solucionadora', 'Não informado')
        po = None
        alcadas = None
        excecao = None
        observacoes = None
        tipo = row.get('Tipo', 'Não informado')
        atendimento_raw = row.get('Atendimento', 'Não informado')
        det_ori_nor = row.get('Determinação / Orientação / Norma', 'Não informado')

        # Mapeamento do cabeçalho (linha 1)
        cabecalho = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
        mapa_colunas = {nome: idx + 1 for idx, nome in enumerate(cabecalho)}

        # Converte os nomes para letras de coluna
        coluna_kb = get_column_letter(mapa_colunas['Base de Conhecimento'])
        coluna_anexo = get_column_letter(mapa_colunas['Anexo'])

        # Acesso às células da linha atual
        linha_excel = i + 2
        celula_kb = ws[f'{coluna_kb}{linha_excel}']
        kbs = celula_kb.hyperlink.target if celula_kb.hyperlink else 'Não têm.'

        celula_anexo = ws[f'{coluna_anexo}{linha_excel}']
        anexo = celula_anexo.hyperlink.target if celula_anexo.hyperlink else 'Não têm.'

        usuario_criador = "Service-Desk/WD"
        revisor_nome = "Service-Desk/WD"
        revisor_email = None
        data_revisao = None
        po_aprovador_nome = None
        po_aprovador_email = None
        data_aprovacao = None
        justificativa_rejeicao = None
        versao = 1.0

        # Dicionário de serviços já criados para evitar duplicata
        chave_servico = f"{nome_subcategoria}::{nome_servico}"

        # EXCEÇÃO: permitir duplicatas para certos serviços
        excecoes_deduplicacao = [
            ("Sistemas Corporativos SSA", "Dúvidas e Orientações"),
        ]

        if chave_servico in servicos_inseridos and nome_subcategoria != "Sistemas Corporativos SSA":
            id_gerado = servicos_inseridos[chave_servico]
        else:
            cursor.execute("SELECT ID FROM subcategoria WHERE TRIM(REPLACE(Titulo, '\n', ' ')) = %s", (nome_subcategoria,))
            resultado = cursor.fetchone()
            if not resultado:
                print(f"[ERRO] Subcategoria não encontrada: {nome_subcategoria}")
                continue

            id_sub = resultado[0]

            cursor.execute("""
                INSERT INTO servico (
                    Titulo, Descricao, tipo, determinacao_orientacao_norma,
                    ID_SubCategoria, UltimaAtualizacao,
                    area_especialista, po_responsavel, alcadas,
                    procedimento_excecao, observacoes, KBs, Anexo, usuario_criador, status_ficha,
                    revisor_nome, revisor_email, data_revisao, po_aprovador_nome,
                    po_aprovador_email, data_aprovacao, justificativa_rejeicao,
                    versao
                )
                VALUES (%s, %s, %s, %s,
                        %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s,
                        %s)
            """, (
                nome_servico, descricao, tipo, det_ori_nor,
                id_sub, agora,
                area, po, alcadas,
                excecao, observacoes, kbs, anexo, usuario_criador, 'rascunho',
                revisor_nome, revisor_email, data_revisao,
                po_aprovador_nome, po_aprovador_email, data_aprovacao,
                justificativa_rejeicao, versao
            ))

            id_gerado = cursor.lastrowid
            codigo_ficha = f"FCH-{id_gerado:04d}"
            cursor.execute("UPDATE servico SET codigo_ficha = %s WHERE ID = %s", (codigo_ficha, id_gerado))
            servicos_inseridos[chave_servico] = id_gerado
            print(f"[OK] Criado servico ID {id_gerado} | Subcategoria: {nome_subcategoria} | Serviço: {nome_servico}")
            
        # ATENDIMENTO
        atendimento_raw = str(row.get('Atendimento') or '')
        atendimentos = [a.strip() for a in re.split(r'[/,\n;\-]', atendimento_raw) if a.strip()][:6]

        descricaotec = row.get('Descrição Técnica')
        descricaotec = descricaotec if pd.notna(descricaotec) else ''
        descricoes_lista = [d.strip() for d in str(descricaotec).split('\n') if d.strip()]

        # Se houver apenas uma descrição, replicar para todos os níveis
        if len(descricoes_lista) == 1 and len(atendimentos) > 1:
            descricoes_lista *= len(atendimentos)
        elif len(descricoes_lista) < len(atendimentos):
            descricoes_lista += ['Descrição técnica padrão.'] * (len(atendimentos) - len(descricoes_lista))

        for idx, nivel in enumerate(atendimentos):
            desc_tecnica = descricoes_lista[idx] if idx < len(descricoes_lista) else 'Descrição técnica padrão.'
            cursor.execute("""
                INSERT INTO servico_atendimento (id_servico, atendimento, descricao_tecnica)
                VALUES (%s, %s, %s)
            """, (id_gerado, nivel, desc_tecnica))


        # SISTEMAS + EQUIPES EXTERNAS (caso especial com blocos múltiplos)
        if aba.strip().lower() in ['sistemas e portais', 'credenciais de acesso', 'segurança', 'infraestrutura', 'microsoft 365', 'softwares homologados', 'portais corporativos']:

            sistemas_raw = str(row.get('Sistema / Portal') or '')
            equipes_raw = str(row.get('Equipe Solucionadora (Externa)', '') or '')

            # Sanitização pesada
            sistemas_raw = re.sub(r'[\u00a0\r\t]', ' ', sistemas_raw)
            equipes_raw = re.sub(r'[\u00a0\r\t]', ' ', equipes_raw)

            # Divide os sistemas por linha e por vírgula/ponto e vírgula
            sistemas = [re.sub(r'\s+', ' ', s).strip() for s in re.split(r'[\n,;]+', sistemas_raw) if s.strip()]
            equipes = [re.sub(r'\s+', ' ', e).strip() for e in re.split(r'[\n,;]+', equipes_raw) if e.strip()]

            # Se só uma equipe, replica para todos os sistemas
            if len(equipes) == 1 and len(sistemas) > 1:
                equipes = [equipes[0]] * len(sistemas)
            elif not equipes:
                equipes = [''] * len(sistemas)
            elif len(equipes) < len(sistemas):
                equipes += [equipes[-1]] * (len(sistemas) - len(equipes))
            elif len(equipes) > len(sistemas):
                equipes = equipes[:len(sistemas)]

            for sistema, equipe in zip(sistemas, equipes):
                cursor.execute("""
                    INSERT INTO servico_sistema (id_servico, nome_sistema)
                    VALUES (%s, %s)
                """, (id_gerado, sistema))
                id_sistema = cursor.lastrowid

                if equipe:
                    cursor.execute("""
                        INSERT INTO servico_equipe_externa (id_servico, nome_equipe, id_sistema)
                        VALUES (%s, %s, %s)
                    """, (id_gerado, equipe, id_sistema))

        # SOFTWARES HOMOLOGADOS
        if 'softwares homologados' in aba.strip().lower():
            software = nome_subcategoria
            versao_software = str(row.get('Versão') or 'Não informada.').strip()[:255]
            if software and versao_software:
                # Buscar todos os IDs de serviço da subcategoria atual
                cursor.execute(
                    "SELECT ID FROM servico WHERE ID_SubCategoria = %s",
                    (id_sub,)
                )
                ids_servicos = [r[0] for r in cursor.fetchall()]
                
                for id_serv in ids_servicos:
                    cursor.execute("""
                        SELECT 1 FROM servico_software
                        WHERE id_servico = %s AND nome_software = %s AND versao_software = %s
                        LIMIT 1
                    """, (id_serv, software, versao_software))
                    existe = cursor.fetchone()
                    
                    if not existe:
                        cursor.execute("""
                            INSERT INTO servico_software (id_servico, nome_software, versao_software)
                            VALUES (%s, %s, %s)
                        """, (id_serv, software, versao_software))

conn.commit()
cursor.close()
conn.close()
print("Serviços importados com sucesso.")